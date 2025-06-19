import sqlite3
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from datetime import datetime, timedelta
from models import (
    UserCreate, UserUpdate, ClassCreate, ClassUpdate, SubjectCreate, SubjectUpdate,
    ThemeCreate, ThemeUpdate, MaterialCreate, MaterialUpdate,
    QuestionOption, QuestionPair, QuestionCreate, FullTestCreate, TestTitleUpdate, TestResultCreate
)
from typing import List, Optional
import openpyxl
import random
import string
from fastapi import Body
from pydantic import BaseModel
import cloudinary
import cloudinary.uploader

cloudinary.config(
    cloud_name='de8bv8y1u',
    api_key='245494152712799',
    api_secret='SfS9q2Z6TtJHHK4e3YivEqCp9i4'
)


DB_PATH = 'brainline.db'

app = FastAPI(
    title="NCFU Brainline API",
    description="Backend для платформы NCFU Brainline",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Авторизация ---
SECRET_KEY = "your_secret_key"  # Замени на свой!
ALGORITHM = "HS256"

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

def generate_login_password(length=8):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choices(chars, k=length))

def authenticate_user(login: str, password: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, login, password, full_name, role FROM users WHERE login = ?", (login,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return None
    user_id, login, db_password, full_name, role = row
    if password != db_password:
        return None
    return {"id": user_id, "login": login, "full_name": full_name, "role": role}

def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    # Не добавляем "exp"!
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=401,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, login, full_name, role FROM users WHERE id = ?", (user_id,))
    row = cursor.fetchone()
    conn.close()
    if row is None:
        raise credentials_exception
    return {"id": row[0], "login": row[1], "full_name": row[2], "role": row[3]}

def is_admin(user):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")

# --- USERS ---
@app.post("/users/", tags=["Пользователи"])
def create_user(data: UserCreate = Body(...), current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Генерируем логин и пароль, если не переданы
    login = getattr(data, "login", None) or generate_login_password()
    password = getattr(data, "password", None) or generate_login_password()
    cursor.execute(
        "INSERT INTO users (full_name, login, password, role) VALUES (?, ?, ?, ?)",
        (data.full_name, login, password, data.role)
    )
    user_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return {
        "id": user_id,
        "full_name": data.full_name,
        "login": login,
        "password": password,
        "role": data.role
    }

@app.get("/users/", tags=["Пользователи"])
def get_users(role: Optional[str] = None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    if role:
        cursor.execute("SELECT id, full_name, login, password, role FROM users WHERE role = ?", (role,))
    else:
        cursor.execute("SELECT id, full_name, login, password, role FROM users")
    users = [{"id": row[0], "full_name": row[1], "login": row[2], "password": row[3], "role": row[4]} for row in cursor.fetchall()]
    conn.close()
    return users

@app.post("/users/bulk_students/", tags=["Пользователи"])
async def bulk_add_students(class_id: int = Form(...), file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    wb = openpyxl.load_workbook(file.file)
    ws = wb.active
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name = row[0]
        if not full_name:
            continue  # пропускаем пустые строки
        login = generate_login_password()
        password = generate_login_password()
        cursor.execute("INSERT INTO users (full_name, login, password, role) VALUES (?, ?, ?, ?)",
                       (full_name, login, password, "student"))
        user_id = cursor.lastrowid
        cursor.execute("INSERT INTO user_classes (user_id, class_id) VALUES (?, ?)", (user_id, class_id))
        count += 1
    conn.commit()
    conn.close()
    return {"count": count}

@app.put("/users/{user_id}", tags=["Пользователи"])
def update_user(user_id: int, data: UserUpdate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET login = ?, password = ?, full_name = ?, role = ? WHERE id = ?",
                   (data.login, data.password, data.full_name, data.role, user_id))
    conn.commit()
    conn.close()
    return {"status": "Пользователь обновлён"}

@app.delete("/users/{user_id}", tags=["Пользователи"])
def delete_user(user_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM user_classes WHERE user_id = ?", (user_id,))
    cursor.execute("DELETE FROM user_subjects WHERE user_id = ?", (user_id,))
    cursor.execute("DELETE FROM test_results WHERE user_id = ?", (user_id,))
    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return {"status": "Пользователь и связанные данные удалены"}

@app.post("/users/bulk_students/", tags=["Пользователи"])
async def bulk_add_students(class_id: int = Form(...), file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    wb = openpyxl.load_workbook(file.file)
    ws = wb.active
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name = row[0]
        if not full_name:
            continue  # пропускаем пустые строки
        login = f"user{str(full_name)[:3].lower()}{count}"
        password = "pass123"
        cursor.execute("INSERT INTO users (full_name, login, password, role) VALUES (?, ?, ?, ?)",
                       (full_name, login, password, "student"))
        user_id = cursor.lastrowid
        cursor.execute("INSERT INTO user_classes (user_id, class_id) VALUES (?, ?)", (user_id, class_id))
        count += 1
    conn.commit()
    conn.close()
    return {"count": count}

@app.delete("/users/bulk_delete/", tags=["Пользователи"])
def bulk_delete_users(start_id: int, end_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users WHERE id BETWEEN ? AND ?", (start_id, end_id))
    deleted = cursor.rowcount
    conn.commit()
    conn.close()
    return {"deleted": deleted}

@app.delete("/users/delete_by_class/", tags=["Пользователи"])
def delete_users_by_class(class_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        DELETE FROM users WHERE id IN (
            SELECT user_id FROM user_classes WHERE class_id = ?
        )
    """, (class_id,))
    deleted = cursor.rowcount
    conn.commit()
    conn.close()
    return {"deleted": deleted}

@app.get("/users/students_without_class/", tags=["Пользователи"])
def get_students_without_class():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, full_name, login, password FROM users
        WHERE role = 'student' AND id NOT IN (
            SELECT user_id FROM user_classes
        )
    """)
    students = [
        {"id": row[0], "full_name": row[1], "login": row[2], "password": row[3]}
        for row in cursor.fetchall()
    ]
    conn.close()
    return students

@app.get("/users/teachers_without_subject/", tags=["Пользователи"])
def get_teachers_without_subject():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, full_name FROM users
        WHERE role = 'teacher' AND id NOT IN (
            SELECT user_id FROM user_subjects
        )
    """)
    teachers = [{"id": row[0], "full_name": row[1]} for row in cursor.fetchall()]
    conn.close()
    return teachers

@app.get("/users/teachers_with_flag/", tags=["Пользователи"])
def get_teachers_with_flag():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, full_name FROM users WHERE role = 'teacher'")
    teachers = []
    for row in cursor.fetchall():
        teacher_id, full_name = row
        cursor.execute("SELECT COUNT(*) FROM user_subjects WHERE user_id = ?", (teacher_id,))
        has_subject = cursor.fetchone()[0] > 0
        teachers.append({
            "id": teacher_id,
            "full_name": full_name,
            "has_subject": has_subject
        })
    conn.close()
    return teachers

@app.get("/users/me", tags=["Пользователи"])
def get_me(current_user: dict = Depends(get_current_user)):
    return current_user
@app.get("/users/{user_id}", tags=["Пользователи"])
def get_user(user_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, full_name, login, role FROM users WHERE id = ?", (user_id,))
    row = cursor.fetchone()
    conn.close()
    if row:
        return {"id": row[0], "full_name": row[1], "login": row[2], "role": row[3]}
    raise HTTPException(status_code=404, detail="User not found")

# --- CLASSES ---
@app.get("/classes/", tags=["Классы"])
def get_classes():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, class_name FROM classes")
    classes = [{"id": row[0], "class_name": row[1]} for row in cursor.fetchall()]
    conn.close()
    return classes

@app.post("/classes/", tags=["Классы"])
def create_class(data: ClassCreate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO classes (class_name) VALUES (?)", (data.class_name,))
    conn.commit()
    class_id = cursor.lastrowid
    conn.close()
    return {"id": class_id, "status": "Класс создан"}

@app.put("/classes/{class_id}", tags=["Классы"])
def update_class(class_id: int, data: ClassUpdate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("UPDATE classes SET class_name = ? WHERE id = ?", (data.class_name, class_id))
    conn.commit()
    conn.close()
    return {"status": "Класс обновлён"}

@app.delete("/classes/{class_id}", tags=["Классы"])
def delete_class(class_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM classes WHERE id = ?", (class_id,))
    conn.commit()
    conn.close()
    return {"status": "Класс удалён"}

@app.get("/classes/{class_id}/students", tags=["Классы"])
def get_class_students(class_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT u.id, u.full_name FROM users u
        JOIN user_classes uc ON u.id = uc.user_id
        WHERE uc.class_id = ?
    """, (class_id,))
    students = [{"id": row[0], "full_name": row[1]} for row in cursor.fetchall()]
    conn.close()
    return students

# --- USER-CLASSES ---
@app.post("/user-classes/", tags=["Классы"])
def add_student_to_class(user_id: int, class_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO user_classes (user_id, class_id) VALUES (?, ?)", (user_id, class_id))
    conn.commit()
    conn.close()
    return {"status": "Ученик добавлен в класс"}

@app.delete("/user-classes/{user_id}", tags=["Классы"])
def remove_student_from_class(user_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM user_classes WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()
    return {"status": "Ученик удалён из класса"}

# --- SUBJECTS ---
@app.get("/subjects/", tags=["Предметы"])
def get_subjects():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, subject_name FROM subjects")
    subjects = [{"id": row[0], "subject_name": row[1]} for row in cursor.fetchall()]
    conn.close()
    return subjects

@app.post("/subjects/", tags=["Предметы"])
def create_subject(data: SubjectCreate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO subjects (subject_name) VALUES (?)", (data.subject_name,))
    conn.commit()
    subject_id = cursor.lastrowid
    conn.close()
    return {"id": subject_id, "status": "Предмет создан"}

@app.put("/subjects/{subject_id}", tags=["Предметы"])
def update_subject(subject_id: int, data: SubjectUpdate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("UPDATE subjects SET subject_name = ? WHERE id = ?", (data.subject_name, subject_id))
    conn.commit()
    conn.close()
    return {"status": "Предмет обновлён"}

@app.delete("/subjects/{subject_id}", tags=["Предметы"])
def delete_subject(subject_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Получаем все темы этого предмета
    cursor.execute("SELECT id FROM themes WHERE subject_id = ?", (subject_id,))
    theme_ids = [row[0] for row in cursor.fetchall()]
    # Удаляем материалы для каждой темы
    for theme_id in theme_ids:
        cursor.execute("DELETE FROM materials WHERE theme_id = ?", (theme_id,))
    # Удаляем темы
    cursor.execute("DELETE FROM themes WHERE subject_id = ?", (subject_id,))
    # Удаляем сам предмет
    cursor.execute("DELETE FROM subjects WHERE id = ?", (subject_id,))
    conn.commit()
    conn.close()
    return {"status": "Предмет, его темы и связанные материалы удалены"}

@app.get("/subjects/{subject_id}/teachers", tags=["Предметы"])
def get_subject_teachers(subject_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT u.id, u.full_name FROM users u
        JOIN user_subjects us ON u.id = us.user_id
        WHERE us.subject_id = ?
    """, (subject_id,))
    teachers = [{"id": row[0], "full_name": row[1]} for row in cursor.fetchall()]
    conn.close()
    return teachers

# --- USER-SUBJECTS ---
@app.post("/user-subjects/", tags=["Предметы"])
def assign_teacher_to_subject(user_id: int, subject_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO user_subjects (user_id, subject_id) VALUES (?, ?)", (user_id, subject_id))
    conn.commit()
    conn.close()
    return {"status": "Учитель назначен на предмет"}

@app.delete("/user-subjects/", tags=["Предметы"])
def remove_teacher_from_subject(user_id: int, subject_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM user_subjects WHERE user_id = ? AND subject_id = ?", (user_id, subject_id))
    conn.commit()
    conn.close()
    return {"status": "Учитель удалён из предмета"}

@app.get("/subjects/with-teachers", tags=["Предметы"])
def get_subjects_with_teachers():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, subject_name FROM subjects")
    subjects_raw = cursor.fetchall()
    subjects = []
    for subject_id, subject_name in subjects_raw:
        cursor2 = conn.cursor()
        cursor2.execute("""
            SELECT u.id, u.full_name
            FROM users u
            JOIN user_subjects us ON u.id = us.user_id
            WHERE us.subject_id = ?
        """, (subject_id,))
        teachers = [{"id": row[0], "full_name": row[1]} for row in cursor2.fetchall()]
        subjects.append({
            "id": subject_id,
            "subject_name": subject_name,
            "teachers": teachers
        })
    conn.close()
    return subjects

# --- THEMES ---
@app.get("/themes/by_subject/{subject_id}", tags=["Темы"])
def get_themes(subject_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, theme_name FROM themes WHERE subject_id = ?", (subject_id,))
    themes = [{"id": row[0], "theme_name": row[1]} for row in cursor.fetchall()]
    conn.close()
    return themes

@app.get("/themes/{theme_id}", tags=["Темы"])
def get_theme(theme_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, theme_name, subject_id FROM themes WHERE id = ?", (theme_id,))
    row = cursor.fetchone()
    conn.close()
    if row:
        return {"id": row[0], "theme_name": row[1], "subject_id": row[2]}
    raise HTTPException(status_code=404, detail="Theme not found")

@app.post("/themes/", tags=["Темы"])
def create_theme(data: ThemeCreate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO themes (theme_name, subject_id) VALUES (?, ?)", (data.theme_name, data.subject_id))
    conn.commit()
    theme_id = cursor.lastrowid
    conn.close()
    return {"id": theme_id, "status": "Тема создана"}

@app.put("/themes/{theme_id}", tags=["Темы"])
def update_theme(theme_id: int, data: ThemeUpdate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("UPDATE themes SET theme_name = ?, subject_id = ? WHERE id = ?", (data.theme_name, data.subject_id, theme_id))
    conn.commit()
    conn.close()
    return {"status": "Тема обновлена"}

@app.delete("/themes/{theme_id}", tags=["Темы"])
def delete_theme(theme_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Сначала удаляем материалы этой темы
    cursor.execute("DELETE FROM materials WHERE theme_id = ?", (theme_id,))
    # Потом удаляем саму тему
    cursor.execute("DELETE FROM themes WHERE id = ?", (theme_id,))
    conn.commit()
    conn.close()
    return {"status": "Тема и связанные материалы удалены"}

# --- MATERIALS ---
@app.get("/materials/by_theme/{theme_id}", tags=["Материалы"])
def get_materials(theme_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, title FROM materials WHERE theme_id = ?", (theme_id,))
    materials = [{"id": row[0], "title": row[1]} for row in cursor.fetchall()]
    conn.close()
    return materials

@app.get("/materials/{material_id}", tags=["Материалы"])
def get_material(material_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, title, markdown_content, theme_id, created_by, created_at FROM materials WHERE id = ?", (material_id,))
    row = cursor.fetchone()
    if row:
        created_by_id = row[4] if len(row) > 4 else None
        created_by_full_name = None
        if created_by_id:
            cursor.execute("SELECT full_name FROM users WHERE id = ?", (created_by_id,))
            user_row = cursor.fetchone()
            if user_row:
                created_by_full_name = user_row[0]
        conn.close()
        return {
            "id": row[0],
            "title": row[1],
            "content": row[2],
            "theme_id": row[3],
            "created_by": created_by_id,
            "created_by_full_name": created_by_full_name,
            "created_at": row[5] if len(row) > 5 else None
        }
    conn.close()
    raise HTTPException(status_code=404, detail="Material not found")

@app.post("/materials/", tags=["Материалы"])
def create_material(data: MaterialCreate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    now = datetime.utcnow().isoformat()
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO materials (title, markdown_content, theme_id, created_by, created_at) VALUES (?, ?, ?, ?, ?)",
        (data.title, data.markdown_content, data.theme_id, current_user["id"], now)
    )
    material_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return {"id": material_id, "status": "Материал создан"}

@app.put("/materials/{material_id}", tags=["Материалы"])
def update_material(material_id: int, data: MaterialUpdate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE materials SET title = ?, markdown_content = ?, theme_id = ? WHERE id = ?",
        (data.title, data.markdown_content, data.theme_id, material_id)
    )
    conn.commit()
    conn.close()
    return {"status": "Материал обновлён"}

@app.delete("/materials/{material_id}", tags=["Материалы"])
def delete_material(material_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM materials WHERE id = ?", (material_id,))
    conn.commit()
    conn.close()
    return {"status": "Материал удалён"}

# --- TESTS ---
@app.get("/tests/by_material/{material_id}", tags=["Тесты"])
def get_tests(material_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, title FROM tests WHERE material_id = ?", (material_id,))
    tests = [{"id": row[0], "title": row[1]} for row in cursor.fetchall()]
    conn.close()
    return tests

@app.get("/tests/{test_id}", tags=["Тесты"])
def get_test(test_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, title FROM tests WHERE id = ?", (test_id,))
    row = cursor.fetchone()
    conn.close()
    if row:
        return {"id": row[0], "title": row[1]}
    raise HTTPException(status_code=404, detail="Test not found")

@app.get("/tests/questions/{test_id}", tags=["Тесты"])
def get_test_questions(test_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, question_text, question_type FROM test_questions WHERE test_id = ?", (test_id,))
    questions = [{"id": row[0], "question_text": row[1], "question_type": row[2]} for row in cursor.fetchall()]
    conn.close()
    return questions

@app.post("/tests/full", tags=["Тесты"])
def create_full_test(test: FullTestCreate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO tests (material_id, title) VALUES (?, ?)",
        (test.material_id, test.title)
    )
    test_id = cursor.lastrowid

    for q in test.questions:
        cursor.execute(
            "INSERT INTO test_questions (test_id, question_text, question_type) VALUES (?, ?, ?)",
            (test_id, q.text, q.type)
        )
        question_id = cursor.lastrowid

        if q.type in ('single', 'multiple'):
            for idx, opt in enumerate(q.options):
                is_correct = (q.type == 'single' and q.correct == idx) or (q.type == 'multiple' and opt.correct)
                cursor.execute(
                    "INSERT INTO test_options (question_id, option_text, is_correct) VALUES (?, ?, ?)",
                    (question_id, opt.text, int(is_correct))
                )
        elif q.type == 'match':
            for pair in q.pairs:
                cursor.execute(
                    "INSERT INTO test_pairs (question_id, left_text, right_text) VALUES (?, ?, ?)",
                    (question_id, pair.left, pair.right)
                )
    conn.commit()
    conn.close()
    return {"id": test_id, "status": "Тест с вопросами создан"}

@app.put("/tests/full/{test_id}", tags=["Тесты"])
def update_full_test(
    test_id: int,
    test: FullTestCreate = Body(...),
    current_user: dict = Depends(get_current_user)
):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Обновить название теста
    cursor.execute("UPDATE tests SET title = ?, material_id = ? WHERE id = ?", (test.title, test.material_id, test_id))
    # Удалить старые вопросы, варианты и пары
    cursor.execute("SELECT id FROM test_questions WHERE test_id = ?", (test_id,))
    question_ids = [row[0] for row in cursor.fetchall()]
    for qid in question_ids:
        cursor.execute("DELETE FROM test_options WHERE question_id = ?", (qid,))
        cursor.execute("DELETE FROM test_pairs WHERE question_id = ?", (qid,))
    cursor.execute("DELETE FROM test_questions WHERE test_id = ?", (test_id,))
    # Добавить новые вопросы
    for q in test.questions:
        cursor.execute(
            "INSERT INTO test_questions (test_id, question_text, question_type) VALUES (?, ?, ?)",
            (test_id, q.text, q.type)
        )
        question_id = cursor.lastrowid
        if q.type in ('single', 'multiple'):
            for idx, opt in enumerate(q.options):
                is_correct = (q.type == 'single' and q.correct == idx) or (q.type == 'multiple' and opt.correct)
                cursor.execute(
                    "INSERT INTO test_options (question_id, option_text, is_correct) VALUES (?, ?, ?)",
                    (question_id, opt.text, int(is_correct))
                )
        elif q.type == 'match':
            for pair in q.pairs:
                cursor.execute(
                    "INSERT INTO test_pairs (question_id, left_text, right_text) VALUES (?, ?, ?)",
                    (question_id, pair.left, pair.right)
                )
    conn.commit()
    conn.close()
    return {"status": "Тест обновлён"}

@app.get("/test_options/{question_id}", tags=["Тесты"])
def get_test_options(question_id: int, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, option_text, is_correct FROM test_options WHERE question_id = ?", (question_id,))
    options = [{"id": row[0], "option_text": row[1], "is_correct": row[2]} for row in cursor.fetchall()]
    conn.close()
    return options

@app.get("/test_pairs/{question_id}", tags=["Тесты"])
def get_test_pairs(question_id: int, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, left_text, right_text FROM test_pairs WHERE question_id = ?", (question_id,))
    pairs = [{"id": row[0], "left_text": row[1], "right_text": row[2]} for row in cursor.fetchall()]
    conn.close()
    return pairs

@app.post("/tests/question/{test_id}", tags=["Тесты"])
def add_question_to_test(test_id: int, q: QuestionCreate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO test_questions (test_id, question_text, question_type) VALUES (?, ?, ?)",
        (test_id, q.text, q.type)
    )
    question_id = cursor.lastrowid

    if q.type in ('single', 'multiple'):
        for idx, opt in enumerate(q.options):
            is_correct = (q.type == 'single' and q.correct == idx) or (q.type == 'multiple' and opt.correct)
            cursor.execute(
                "INSERT INTO test_options (question_id, option_text, is_correct) VALUES (?, ?, ?)",
                (question_id, opt.text, int(is_correct))
            )
    elif q.type == 'match':
        for pair in q.pairs:
            cursor.execute(
                "INSERT INTO test_pairs (question_id, left_text, right_text) VALUES (?, ?, ?)",
                (question_id, pair.left, pair.right)
            )
    conn.commit()
    conn.close()
    return {"id": question_id, "status": "Вопрос добавлен"}

@app.delete("/tests/question/{question_id}", tags=["Тесты"])
def delete_question(question_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM test_questions WHERE id = ?", (question_id,))
    conn.commit()
    conn.close()
    return {"status": "Вопрос удалён"}

@app.delete("/tests/{test_id}", tags=["Тесты"])
def delete_test(test_id: int, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Получаем все вопросы этого теста
    cursor.execute("SELECT id FROM test_questions WHERE test_id = ?", (test_id,))
    question_ids = [row[0] for row in cursor.fetchall()]
    # Удаляем все варианты и пары для этих вопросов
    for qid in question_ids:
        cursor.execute("DELETE FROM test_options WHERE question_id = ?", (qid,))
        cursor.execute("DELETE FROM test_pairs WHERE question_id = ?", (qid,))
    # Удаляем вопросы
    cursor.execute("DELETE FROM test_questions WHERE test_id = ?", (test_id,))
    # Удаляем сам тест
    cursor.execute("DELETE FROM tests WHERE id = ?", (test_id,))
    conn.commit()
    conn.close()
    return {"status": "Тест и связанные вопросы удалены"}

@app.put("/tests/{test_id}", tags=["Тесты"])
def update_test_title(test_id: int, data: TestTitleUpdate, current_user: dict = Depends(get_current_user)):
    is_admin(current_user)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("UPDATE tests SET title = ? WHERE id = ?", (data.title, test_id))
    conn.commit()
    conn.close()
    return {"status": "Название теста обновлено"}

@app.get("/tests/", tags=["Тесты"])
def get_all_tests(current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, title FROM tests")
    tests = [{"id": row[0], "title": row[1]} for row in cursor.fetchall()]
    conn.close()
    return tests

@app.get("/subjects/with_tests", tags=["Предметы"])
def get_subjects_with_tests(current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT s.id, s.subject_name, MAX(tst.id) as max_test_id
        FROM subjects s
        JOIN themes t ON t.subject_id = s.id
        JOIN materials m ON m.theme_id = t.id
        JOIN tests tst ON tst.material_id = m.id
        GROUP BY s.id, s.subject_name
        ORDER BY max_test_id DESC
    """)
    subjects = [{"id": row[0], "subject_name": row[1], "max_test_id": row[2]} for row in cursor.fetchall()]
    conn.close()
    return subjects

@app.post("/test_results/", tags=["Тесты"])
def save_test_result(data: TestResultCreate, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    now = datetime.utcnow().isoformat()
    # Проверяем, есть ли уже результат
    cursor.execute("""
        SELECT id FROM test_results WHERE user_id = ? AND test_id = ?
    """, (current_user["id"], data.test_id))
    row = cursor.fetchone()
    if row:
        # Обновляем существующий результат
        cursor.execute("""
            UPDATE test_results
            SET correct_answers = ?, total_questions = ?, percent = ?, passed_at = ?
            WHERE id = ?
        """, (data.correct_answers, data.total_questions, data.percent, now, row[0]))
    else:
        # Вставляем новый результат
        cursor.execute("""
            INSERT INTO test_results (user_id, test_id, correct_answers, total_questions, percent, passed_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (current_user["id"], data.test_id, data.correct_answers, data.total_questions, data.percent, now))
    conn.commit()
    conn.close()
    return {"status": "Результат сохранён"}

@app.get("/test_results/my")
def get_my_results(current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT tr.id, t.title, m.title, th.theme_name, s.subject_name,
               tr.correct_answers, tr.total_questions, tr.percent, tr.passed_at
        FROM test_results tr
        JOIN tests t ON tr.test_id = t.id
        JOIN materials m ON t.material_id = m.id
        JOIN themes th ON m.theme_id = th.id
        JOIN subjects s ON th.subject_id = s.id
        WHERE tr.user_id = ?
        ORDER BY tr.passed_at DESC
    """, (current_user["id"],))
    results = [
        {
            "id": row[0],
            "test_title": row[1],
            "material_title": row[2],
            "theme_name": row[3],
            "subject_name": row[4],
            "correct_answers": row[5],
            "total_questions": row[6],
            "percent": row[7],
            "passed_at": row[8]
        }
        for row in cursor.fetchall()
    ]
    conn.close()
    return results

@app.get("/themes/with_tests/{subject_id}", tags=["Темы"])
def get_themes_with_tests(subject_id: int, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT t.id, t.theme_name, MAX(tst.id) as max_test_id
        FROM themes t
        JOIN materials m ON m.theme_id = t.id
        JOIN tests tst ON tst.material_id = m.id
        WHERE t.subject_id = ?
        GROUP BY t.id, t.theme_name
        ORDER BY max_test_id DESC
    """, (subject_id,))
    themes = [{"id": row[0], "theme_name": row[1], "max_test_id": row[2]} for row in cursor.fetchall()]
    conn.close()
    return themes

@app.get("/materials/with_tests/{theme_id}", tags=["Материалы"])
def get_materials_with_tests(theme_id: int, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT m.id, m.title, tst.id as test_id, MAX(tst.id) as max_test_id
        FROM materials m
        JOIN tests tst ON tst.material_id = m.id
        WHERE m.theme_id = ?
        GROUP BY m.id, m.title
        ORDER BY max_test_id DESC
    """, (theme_id,))
    materials = [
        {"id": row[0], "title": row[1], "test_id": row[2], "max_test_id": row[3]}
        for row in cursor.fetchall()
    ]
    conn.close()
    return materials

@app.post("/token", tags=["Авторизация"])
def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends()):
    user = authenticate_user(form_data.username, form_data.password)
    if not user:
        raise HTTPException(status_code=400, detail="Incorrect username or password")
    access_token = create_access_token(
        data={"sub": str(user["id"])}
    )
    # Для фронта удобно вернуть роль сразу
    return {"access_token": access_token, "token_type": "bearer", "role": user["role"]}


@app.get("/search/all")
def search_all(q: str, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    result = []
    # Предметы
    cursor.execute("SELECT id, subject_name FROM subjects WHERE subject_name LIKE ?", (f"%{q}%",))
    for row in cursor.fetchall():
        result.append({"id": row[0], "title": row[1], "type": "subject"})
    # Темы
    cursor.execute("SELECT id, theme_name FROM themes WHERE theme_name LIKE ?", (f"%{q}%",))
    for row in cursor.fetchall():
        result.append({"id": row[0], "title": row[1], "type": "theme"})
    # Материалы
    cursor.execute("SELECT id, title FROM materials WHERE title LIKE ?", (f"%{q}%",))
    for row in cursor.fetchall():
        result.append({"id": row[0], "title": row[1], "type": "material"})
    # Тесты
    cursor.execute("SELECT id, title FROM tests WHERE title LIKE ?", (f"%{q}%",))
    for row in cursor.fetchall():
        result.append({"id": row[0], "title": row[1], "type": "test"})
    # Поиск по медиа
    cursor.execute("""
        SELECT id, url, material_id FROM media
        WHERE url LIKE ? OR public_id LIKE ?
    """, (f"%{q}%", f"%{q}%"))
    for row in cursor.fetchall():
        result.append({
            "type": "media",
            "id": row[0],
            "title": row[1].split('/')[-1],  # имя файла
            "material_id": row[2]
        })
    # Поиск по медиа по названию материала
    cursor.execute("""
        SELECT media.id, media.url, media.material_id, materials.title
        FROM media
        JOIN materials ON media.material_id = materials.id
        WHERE materials.title LIKE ?
    """, (f"%{q}%",))
    for row in cursor.fetchall():
        result.append({
            "type": "media",
            "id": row[0],
            "title": row[3],  # название материала
            "material_id": row[2]
        })
    conn.close()
    return result

@app.get("/test_results/my")
def get_my_results(current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT tr.id, t.title, m.title, th.theme_name, s.subject_name,
               tr.correct_answers, tr.total_questions, tr.percent, tr.passed_at
        FROM test_results tr
        JOIN tests t ON tr.test_id = t.id
        JOIN materials m ON t.material_id = m.id
        JOIN themes th ON m.theme_id = th.id
        JOIN subjects s ON th.subject_id = s.id
        WHERE tr.user_id = ?
        ORDER BY tr.passed_at DESC
    """, (current_user["id"],))
    results = [
        {
            "id": row[0],
            "test_title": row[1],
            "material_title": row[2],
            "theme_name": row[3],
            "subject_name": row[4],
            "correct_answers": row[5],
            "total_questions": row[6],
            "percent": row[7],
            "passed_at": row[8]
        }
        for row in cursor.fetchall()
    ]
    conn.close()
    return results

@app.get("/test_results/class")
def get_class_results(class_id: int, material_id: int, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT u.id, u.full_name, 
            COALESCE(MAX(tr.percent), 0),
            COALESCE(MAX(tr.correct_answers), 0),
            COALESCE(MAX(tr.total_questions), 0)
        FROM users u
        JOIN user_classes uc ON u.id = uc.user_id
        LEFT JOIN test_results tr ON tr.user_id = u.id
        LEFT JOIN tests t ON tr.test_id = t.id
        WHERE uc.class_id = ? AND t.material_id = ?
        GROUP BY u.id, u.full_name
        ORDER BY u.full_name
    """, (class_id, material_id))
    results = [
        {
            "user_id": row[0],
            "full_name": row[1],
            "percent": row[2],
            "correct_answers": row[3],
            "total_questions": row[4]
        }
        for row in cursor.fetchall()
    ]
    conn.close()
    return results

@app.post("/media/upload/", tags=["Медиа"])
async def upload_media(
    material_id: int = Form(...),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    # Только учитель или админ
    if current_user["role"] not in ("admin", "teacher"):
        raise HTTPException(status_code=403, detail="Нет доступа")
    # Загружаем файл на Cloudinary
    if file.content_type.startswith("video/"):
        resource_type = "video"
    else:
        resource_type = "auto"
    result = cloudinary.uploader.upload(
        file.file,
        resource_type=resource_type
    )
    url = result["secure_url"]
    public_id = result["public_id"]
    media_type = result["resource_type"]  # image, video и т.д.
    uploaded_at = datetime.utcnow().isoformat()

    # Сохраняем в БД
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO media (material_id, url, type, public_id, uploaded_at) VALUES (?, ?, ?, ?, ?)",
        (material_id, url, media_type, public_id, uploaded_at)
    )
    conn.commit()
    conn.close()
    return {"url": url, "type": media_type, "public_id": public_id}

@app.get("/materials/", tags=["Материалы"])
def get_all_materials(current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, title FROM materials")
    materials = [{"id": row[0], "title": row[1]} for row in cursor.fetchall()]
    conn.close()
    return materials

@app.get("/media/by_material/{material_id}", tags=["Медиа"])
def get_media_by_material(material_id: int, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, url, type, public_id, uploaded_at FROM media WHERE material_id = ?", (material_id,))
    media = [
        {
            "id": row[0],
            "url": row[1],
            "type": row[2],
            "public_id": row[3],
            "uploaded_at": row[4]
        }
        for row in cursor.fetchall()
    ]
    conn.close()
    return media


@app.delete("/media/{media_id}", tags=["Медиа"])
def delete_media(media_id: int, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ("admin", "teacher"):
        raise HTTPException(status_code=403, detail="Нет доступа")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT public_id, type FROM media WHERE id = ?", (media_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Медиафайл не найден")
    public_id, media_type = row
    # Удаляем из Cloudinary
    if public_id:
        cloudinary.uploader.destroy(public_id, invalidate=True, resource_type=media_type)
    # Удаляем из БД
    cursor.execute("DELETE FROM media WHERE id = ?", (media_id,))
    conn.commit()
    conn.close()
    return {"status": "Медиафайл удалён"}

@app.get("/subjects/with_media")
def get_subjects_with_media(current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT s.id, s.subject_name, MAX(med.id) as max_media_id
        FROM subjects s
        JOIN themes t ON t.subject_id = s.id
        JOIN materials mat ON mat.theme_id = t.id
        JOIN media med ON med.material_id = mat.id
        GROUP BY s.id, s.subject_name
    """)
    result = [
        {"id": row[0], "subject_name": row[1], "max_media_id": row[2]}
        for row in cursor.fetchall()
    ]
    conn.close()
    return result

@app.get("/themes/with_media/{subject_id}", tags=["Темы"])
def get_themes_with_media(subject_id: int, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT t.id, t.theme_name, MAX(med.id) as max_media_id
        FROM themes t
        JOIN materials m ON m.theme_id = t.id
        JOIN media med ON med.material_id = m.id
        WHERE t.subject_id = ?
        GROUP BY t.id, t.theme_name
        ORDER BY max_media_id DESC
    """, (subject_id,))
    themes = [{"id": row[0], "theme_name": row[1], "max_media_id": row[2]} for row in cursor.fetchall()]
    conn.close()
    return themes

@app.get("/materials/with_media/{theme_id}", tags=["Материалы"])
def get_materials_with_media(theme_id: int, current_user: dict = Depends(get_current_user)):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT m.id, m.title, MAX(media.id) as max_media_id
        FROM materials m
        JOIN media ON media.material_id = m.id
        WHERE m.theme_id = ?
        GROUP BY m.id, m.title
        ORDER BY max_media_id DESC
    """, (theme_id,))
    materials = [{"id": row[0], "title": row[1], "max_media_id": row[2]} for row in cursor.fetchall()]
    conn.close()
    return materials